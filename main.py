import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import json
import os.path
import gzip
import shutil
from openpyxl.styles import Font, Border, Side


def unzip_file(from_file, to_file):
    """Extract .tsv file from .gz file."""
    with gzip.open(from_file, 'rb') as f_in:
        with open(to_file, 'wb') as f_out:
            shutil.copyfileobj(f_in, f_out)
    print("File unzip to %s successfully." % to_file)


def check_file():
    """
    Check for the existence of a database file or archive in the project directory.
    Download and unzip the archive.
    """
    db_file_name = 'data.tsv'
    zip_db_file = 'name.basics.tsv.gz'
    script_dir = os.path.dirname(os.path.abspath(__file__))

    db_file_full_path = os.path.join(script_dir, db_file_name)
    zip_db_file_full_path = os.path.join(script_dir, zip_db_file)

    if os.path.isfile(db_file_full_path):
        print('Database file %s found.' % db_file_name)
    else:
        if os.path.isfile(zip_db_file_full_path):
            print('Database zip file %s found.' % zip_db_file)
            unzip_file(zip_db_file_full_path, db_file_full_path)
        else:
            print('Database file %s not found.\nDatabase zip file %s found.\n'
                  'File download starts...' % (db_file_name, zip_db_file))
            r = requests.get('https://datasets.imdbws.com/name.basics.tsv.gz', stream=True)
            zip_db_file_full_path = os.path.join(script_dir, zip_db_file)
            with open(os.path.join(script_dir, zip_db_file), 'wb') as fd:
                for chunk in r.iter_content(chunk_size=128):
                    fd.write(chunk)
                print("File %s downloaded successfully." % zip_db_file)
            unzip_file(zip_db_file_full_path, db_file_full_path)


def read_all_id_from_database():
    """Read database file and return list of ID."""
    with open('data.tsv', 'r', encoding='utf-8') as db_file:
        id_list = [line.split()[0] for line in db_file]
    return id_list


def get_html(url):
    """Return response of get request."""
    response = (requests.get(url, headers={'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.5, '
                                                     'image/webp,*/*;q=0.5',
                                           'Accept-Language': 'en-US;q=0.5,en;q=0.3'}))
    return response


def get_content(html):
    """
    Parsing HTML page with BeautifulSoup.
    :param html: content of the server’s response.
    :return: dictionary with actor's name and his filmography.
    """
    main_url = 'https://www.imdb.com'
    filmography = {}
    actor_full_name = ""

    soup = BeautifulSoup(html, 'html.parser')

    items = soup.find('div', attrs={'id': 'filmography'}).find_all_next('div', class_='filmo-category-section')
    categories = soup.find('div', attrs={'id': 'filmography'}).find_all_next('div', class_='head')
    actor_name = soup.find('table', attrs={'id': 'name-overview-widget-layout'}).find('h1').get_text().split('\n')[:-1]

    for i in actor_name:
        actor_full_name += i + ' '
    actor_full_name = actor_full_name[1:len(actor_full_name)-1]

    for category_raw in categories:
        category = category_raw['data-category']
        filmography[category] = {'count': 0, 'films': []}

    for item in items:
        films = item.find_all('div', class_='filmo-row')
        for film in films:
            film_id = film['id']
            category_id = film_id[:film_id.find('-')]
            film_name = film.find('b').get_text()
            film_year = film.find('span', class_='year_column').get_text().replace('\n', '').split()
            if not film_year:
                film_year = "No Year"
            else:
                film_year = film_year[0]
            film_link = main_url + film.find('b').a['href']
            filmography[category_id]['count'] += 1
            filmography[category_id]['films'].append({'title': film_name, 'link': film_link, 'year': film_year})

    return {'actor_full_name': actor_full_name, 'filmography': filmography}


def create_json(filmography, actor_name):
    """
    Save actor's filmography to JSON file.
    :param filmography: dictionary with actor's filmography.
    :param actor_name: actor's name.
    :return: create JSON file.
    """
    file_name = actor_name.replace(" ", "_") + '.json'
    with open(file_name, 'w') as f:
        json.dump(filmography, f, indent=4)


def save_to_excel(filmography, actor_name):
    """
    Save actor's filmography to XLSX file.
    :param filmography: dictionary with actor's filmography.
    :param actor_name: actor's name.
    :return: create XLSX file.
    """

    wb = Workbook()

    for category in filmography.keys():
        ws = wb.create_sheet(category)
        ws['A1'] = 'Title'
        ws['B1'] = 'Link'
        ws['C1'] = 'Year'

        medium_border = (Border(left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium')))

        thin_border = (Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin')))

        headers = ws[1]
        for cell in headers:
            cell.font = Font(bold=True)
            cell.border = medium_border

        for i, item in enumerate(filmography[category]['films']):
            _ = ws.cell(row=i+2, column=1, value=item['title'])
            _ = ws.cell(row=i+2, column=2, value=item['link'])
            _ = ws.cell(row=i+2, column=3, value=item['year'])

        for row in ws.iter_rows(min_row=2, max_col=3):
            for cell in row:
                cell.border = thin_border

    wb.remove(wb['Sheet'])
    file_name = actor_name + '.xlsx'
    wb.save(file_name)


def main():

    check_file()

    id_list = read_all_id_from_database()
    url = 'https://www.imdb.com/name/'

    while True:
        actor_id = input('Enter the actor\'s id from IMDB database: ')
        if actor_id not in id_list:
            print('Actor with ID = ' + actor_id + ' not found in the database. Try another ID')
        else:
            break

    full_url = url + actor_id + '/'
    html = get_html(full_url)
    content = get_content(html.text)

    actor_full_name = content['actor_full_name']
    filmography = content['filmography']

    print('Your actor is', actor_full_name)
    print('Filmography: ')

    for i in filmography.keys():
        print(i, ': ', filmography[i]['count'], sep="")

    create_json(filmography, actor_full_name)

    print('Full information saved to', actor_full_name.replace(" ", "_") + '.json and to '
          + actor_full_name.replace(" ", "_") + '.xlsx')

    save_to_excel(filmography, actor_full_name)


if __name__ == '__main__':
    main()
